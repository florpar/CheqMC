# -*- coding: utf-8 -*-
"""
Created on Thu Jun  5 13:39:52 2025

@author: Usuario
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
from pathlib import Path
import re
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageTk

path_yag = "C:\\Yaguarete\\Standards"
NUCLIDE_COL = 'Nucleido'

# Función para normalizar nombres de nucleidos
def normalize_nuclide_name(nuclide_name: str) -> str:
    try:
        nuclide_name = str(nuclide_name).replace('-', '')
        match = re.match(r'^([A-Za-z]+)(\d+)?([a-zA-Z]*)$', nuclide_name)
        if not match:
            return nuclide_name
        element_symbol, mass_number, metastable_state = match.groups()
        return f"{element_symbol.capitalize()}{mass_number or ''}{metastable_state.lower()}"
    except Exception:
        return nuclide_name

# Leer archivo codificacion.xlsx
def cargar_codificacion(carpeta_base):
    try:
        cod_path = os.path.join(carpeta_base, 'codificacion.xlsx')
        df = pd.read_excel(cod_path, names=['sname', 'cert_file', 'humedad'])
        return df
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el archivo codificacion.xlsx:\n{e}")
        return pd.DataFrame()

# Leer archivo .txt del cert_file
def read_certificate_data(cert_path):
    try:
        cert_df = pd.read_csv(cert_path, sep='\\s+', header=None, names=[NUCLIDE_COL, 'C_standard', 'delta_C_standard'])
        cert_df[NUCLIDE_COL] = cert_df[NUCLIDE_COL].apply(normalize_nuclide_name)
        cert_df = cert_df.drop_duplicates(subset=[NUCLIDE_COL], keep='first')
        cert_df['C_standard'] = pd.to_numeric(cert_df['C_standard'], errors='coerce')
        cert_df['delta_C_standard'] = pd.to_numeric(cert_df['delta_C_standard'], errors='coerce')
        return cert_df
    except Exception as e:
        print(f"Error reading cert file: {e}")
        return pd.DataFrame()

# Extraer columnas con FM Corr, Inc Corr, LD Corr para material dado
def extract_multiindex_result_safe(file_path, material):
    try:
        df_raw = pd.read_excel(file_path, sheet_name="Mediciones Corregidas", header=None)
        header_row_idx = df_raw.index[df_raw.iloc[:, 0] == 'Nucleido'].tolist()[0]
        df_main = df_raw.iloc[header_row_idx + 1:].copy()
        df_main.columns = [f"col_{i}" for i in range(df_raw.shape[1])]  # safe headers
        for col_idx, label in enumerate(['Nucleido', 'Standard', 'Tipo', 'Energia', 'Detector']):
            df_main[label] = df_raw.iloc[header_row_idx + 1:, col_idx].ffill()
        row0 = df_raw.iloc[0]
        row1 = df_raw.iloc[1]
        archivo = Path(file_path).stem.replace("_resultados", "")
        suffixes = ['FM Corr', 'Inc Corr']
        result_parts = []
        for i in range(len(row0) - 1):
            if str(row0[i]).strip().upper() == material.upper():
                sublabels = [str(row1[i]).strip(), str(row1[i+1]).strip()]
                if sublabels == suffixes:
                    part = df_main[[f'col_{i}', f'col_{i+1}','Nucleido', 'Standard', 'Tipo', 'Energia', 'Detector']].copy()
                    part.columns = [f"{archivo} FM", f"{archivo} Inc",'Nucleido', 'Standard', 'Tipo', 'Energia', 'Detector']
                    part = part.set_index(['Nucleido', 'Standard', 'Tipo', 'Energia', 'Detector'])
                    result_parts.append(part)
        if result_parts:
            result_df = pd.concat(result_parts).groupby(level=[0,1,2,3,4]).first()
            return result_df
        else:
            return None
    except Exception as e:
        print(f"Error procesando {file_path}: {e}")
        return None

def group_energies(df: pd.DataFrame) -> pd.DataFrame:
    try:
        df = df.copy()
        grouped_energies = {}
        for nuclide in df['Nucleido']:
            energies = df[df['Nucleido'] == nuclide]['Energia'].unique()
            energies.sort()
            energy_groups = []
            for energy in energies:
                assigned = False
                for group in energy_groups:
                    representative_energy = group['representative']
                    if abs(energy - representative_energy) <= 0.05 * representative_energy:
                        group['energies'].append(energy)
                        assigned = True
                        break
                if not assigned:
                    energy_groups.append({'energies': [energy], 'representative': energy})
            # Map energies to representatives
            for group in energy_groups:
                representative_energy = group['representative']
                for energy in group['energies']:
                    grouped_energies[(nuclide, energy)] = representative_energy
        # Apply grouping
        df['Energia'] = df.apply(lambda row: grouped_energies.get((row['Nucleido'], row['Energia'])), axis=1)
        return df
    except Exception as e:
        print(f"Error grouping energies: {str(e)}")
        return df

def process_all_files_multiindex(material, folder_path, output_path):
    dfs = []
    for file in os.listdir(folder_path):
        if file.endswith("_resultados.xlsx"):
            df = extract_multiindex_result_safe(os.path.join(folder_path, file), material)
            if df is not None:
                dfs.append(df)
    if dfs:
        final_df = pd.concat(dfs, axis=1).reset_index()
        # 💡 Agrupar energías
        final_df = group_energies(final_df)
        # 💡 Agrupar por campos clave para consolidar en una fila
        final_df = final_df.groupby(["Nucleido", "Standard", "Tipo", "Energia", "Detector"], as_index=False).first()
        final_df = final_df.sort_values(by=["Nucleido", "Standard", "Detector"])
        return final_df
    return pd.DataFrame()

def apply_red_format_if_out_of_range(df, cert_df, output_path):
    wb = Workbook()
    ws = wb.active
    cert_dict = cert_df.set_index(NUCLIDE_COL).to_dict('index')
    fm_cols = [c for c in df.columns if "FM" in c]
    column_order = []
    for c in df.columns:
        column_order.append(c)  # primero incluimos todo tal cual
    for fm_col in fm_cols:
        inc_col = fm_col.replace("FM", "Inc")
        z_col = fm_col.replace("FM", "Z-Score")
        df[z_col] = None
    new_cols = []
    for col in df.columns:
        if col.endswith(" FM"):
            inc_col = col.replace("FM", "Inc")
            z_col = col.replace("FM", "Z-Score")
            new_cols.extend([col, inc_col, z_col])
        elif col.endswith("Inc"):
            continue  # ya agregado antes junto con FM
        elif col.endswith("Z-Score"):
            continue  # ya agregado junto con FM
        elif col not in new_cols:
            new_cols.append(col)
    df = df[new_cols]  # aplicar nuevo orden
    df.insert(df.columns.get_loc('Detector') + 1, 'Cert FM', df[NUCLIDE_COL].map(lambda x: cert_dict.get(normalize_nuclide_name(x), {}).get('C_standard')))
    df.insert(df.columns.get_loc('Cert FM') + 1, 'Cert Inc', df[NUCLIDE_COL].map(lambda x: cert_dict.get(normalize_nuclide_name(x), {}).get('delta_C_standard')))
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    cert_dict = cert_df.set_index(NUCLIDE_COL).to_dict('index')
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    nuclide_candidates = {}
    nucleido_col_idx = df.columns.get_loc('Nucleido') + 1
    for fm_col in fm_cols:
        inc_col = fm_col.replace("FM", "Inc")
        z_col = fm_col.replace("FM", "Z-Score")
        fm_idx = df.columns.get_loc(fm_col) + 1
        inc_idx = df.columns.get_loc(inc_col) + 1
        z_idx = df.columns.get_loc(z_col) + 1
        for row_idx in range(2, len(df) + 2):
            nuclide = normalize_nuclide_name(ws.cell(row=row_idx, column=nucleido_col_idx).value)
            if nuclide not in cert_dict:
                continue
            c_std = cert_dict[nuclide]['C_standard']
            d_std = cert_dict[nuclide]['delta_C_standard']
            if pd.isna(c_std) or pd.isna(d_std):
                continue
            try:
                fm_val = float(ws.cell(row=row_idx, column=fm_idx).value)
                inc_val = float(ws.cell(row=row_idx, column=inc_idx).value)
                lower_corr = fm_val - inc_val
                upper_corr = fm_val + inc_val
                lower_cert = c_std - d_std
                upper_cert = c_std + d_std
                extended_lower = c_std - 3 * d_std
                extended_upper = c_std + 3 * d_std
                z_score = abs(fm_val - c_std) / ((inc_val**2 + d_std**2)**0.5)
                ws.cell(row=row_idx, column=z_idx, value=z_score)
                df.at[row_idx-2, z_col] = z_score
                if z_score > 2:
                    ws.cell(row=row_idx, column=fm_idx).fill = red_fill
                    ws.cell(row=row_idx, column=inc_idx).fill = red_fill
                    ws.cell(row=row_idx, column=z_idx).fill = red_fill
                # Chequear si NO hay intersección entre intervalos
                elif upper_corr < lower_cert or upper_cert < lower_corr:
                    # Fuera de rango, pintar amarillo
                    ws.cell(row=row_idx, column=fm_idx).fill = yellow_fill
                    ws.cell(row=row_idx, column=inc_idx).fill = yellow_fill
                elif extended_lower <= fm_val <= extended_upper:
                    diff = abs(fm_val - c_std)
                    if nuclide not in nuclide_candidates or diff < nuclide_candidates[nuclide][2]:
                        nuclide_candidates[nuclide] = (row_idx, fm_idx, inc_idx, diff)
            except:
                continue
    for r_idx, fm_col, inc_col, _ in nuclide_candidates.values():
        ws.cell(row=r_idx, column=fm_col).fill = green_fill
        ws.cell(row=r_idx, column=fm_col).font = bold_font
        ws.cell(row=r_idx, column=inc_col).fill = green_fill
        ws.cell(row=r_idx, column=inc_col).font = bold_font

    wb.save(output_path)

# Interfaz Gráfica
def run_gui():
    root = tk.Tk()
    root.title("Chequeo Material Control")
    root.geometry("600x250")
    logo_path = "C:\Yaguarete\CheqMC\CMC.ico"  
    img = Image.open(logo_path)
    icon_image = ImageTk.PhotoImage(img)
    root.iconphoto(False, icon_image)  #
    codificacion_df = cargar_codificacion(path_yag)
    snames = codificacion_df['sname'].dropna().unique().tolist()
    selected_sname = tk.StringVar()
    selected_cert = tk.StringVar()
    folder_var = tk.StringVar()
    # Función que actualiza los valores del toggle de certificados
    def update_cert_file(event):
        sname = selected_sname.get()
        cert= codificacion_df[codificacion_df['sname'] == sname]['cert_file']
        txt_files = [f for f in os.listdir(path_yag) if f.endswith(".txt")]
        cert_combobox['values'] = txt_files
        if not cert.empty:
            sugg_co = cert.iloc[0]
            suggested = f"{sugg_co}.txt"
            cert_combobox.set(suggested if suggested in txt_files else "Seleccionar archivo")
        else:
            cert_combobox.set("Seleccionar archivo")
    def browse_folder():
        folder = filedialog.askdirectory()
        folder_var.set(folder)
    def generate():
        material = selected_sname.get().strip()
        folder_path = folder_var.get()
        cert_file = selected_cert.get()
        if not material:
            messagebox.showerror("Error", "Seleccioná un material de control.")
            return
        if not folder_path:
            messagebox.showerror("Error", "Seleccioná una carpeta con archivos.")
            return
        if not cert_file or not cert_file.endswith('.txt'):
            messagebox.showerror("Error", "Seleccioná un archivo certificado válido (.txt).")
            return
        # Pide nombre base
        output_name = filedialog.asksaveasfilename(
            defaultextension="",
            filetypes=[("Nombre base del archivo", "*")],
            title="Escribí el nombre base para los archivos")
        if not output_name:
            return
        output_dir = os.path.dirname(output_name)
        base_name = os.path.basename(output_name)  # solo el nombre que escribió el usuario
        output_base = os.path.join(output_dir, f"{base_name}_control_material")
        os.makedirs(output_base, exist_ok=True)
        path_base = os.path.join(output_base, f"{base_name}_{material}_control.xlsx")
        path_rangos = os.path.join(output_base, f"{base_name}_{material}_rangos.xlsx")
        cert_path = os.path.join(path_yag, cert_file)
        df_fin = process_all_files_multiindex(material, folder_path, path_base)
        if df_fin.empty:
            messagebox.showwarning("Sin datos", "No se generaron resultados para ese material.")
            return
        df_fin.to_excel(path_base, index=False)
        cert_df = read_certificate_data(cert_path)
        apply_red_format_if_out_of_range(df_fin, cert_df, path_rangos)
        messagebox.showinfo("Éxito",f"Archivos generados correctamente:\n{path_base}\n{path_rangos}")
        os.startfile(path_rangos)
    # Interfaz gráfica
    tk.Label(root, text="Seleccioná el material :").grid(row=0, column=0, sticky='w', padx=10, pady=5)
    combo = ttk.Combobox(root, textvariable=selected_sname, values=snames, width=40)
    combo.grid(row=0, column=1, padx=10, pady=5)
    combo.bind("<<ComboboxSelected>>", update_cert_file)
    tk.Label(root, text="Archivo certificado :").grid(row=1, column=0, sticky='w', padx=10, pady=5)
    cert_combobox = ttk.Combobox(root, textvariable=selected_cert, width=37)
    cert_combobox.grid(row=1, column=1, padx=10, pady=5)
    tk.Label(root, text="Carpeta con archivos:").grid(row=2, column=0, sticky='w', padx=10, pady=5)
    tk.Entry(root, textvariable=folder_var, width=40).grid(row=2, column=1, padx=10, pady=5)
    tk.Button(root, text="Buscar", command=browse_folder).grid(row=2, column=2)
    tk.Button(root, text="Generar Comparativo", command=generate).grid(row=4, column=0, columnspan=3, pady=20)
    
    root.mainloop()

if __name__ == "__main__":
    run_gui()
